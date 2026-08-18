"""Excel parser. Reads native cell values per sheet, detects table regions and
headers (which may start on any row), and supports sheet selection.
"""
from __future__ import annotations

import io
from typing import List, Optional

from ..extraction.table_detection import tables_from_grid
from ..models.documents import (ExtractionResult, FinancialDocument, FinancialTable,
                                 ParseWarning)
from ..models.enums import FileFormat, Severity, WarningCode
from ..utils.errors import ParserError

try:
    import openpyxl
    _OPENPYXL = True
except Exception:  # pragma: no cover
    _OPENPYXL = False


def list_sheets(filename: str, data: bytes) -> List[str]:
    """Sheet names without a full parse (for a UI sheet picker)."""
    if filename.lower().endswith(".xls"):
        import pandas as pd
        return list(pd.read_excel(io.BytesIO(data), sheet_name=None, header=None).keys())
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def _sheet_grids(filename: str, data: bytes):
    """Yield (sheet_name, grid) using native types, expanding merged cells."""
    if filename.lower().endswith(".xls"):
        import pandas as pd
        sheets = pd.read_excel(io.BytesIO(data), sheet_name=None, header=None, dtype=object)
        for name, df in sheets.items():
            yield name, [list(r) for r in df.itertuples(index=False, name=None)]
        return
    # data_only=True gives computed values; not read_only so merges are available
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    try:
        for ws in wb.worksheets:
            grid = [list(row) for row in ws.iter_rows(values_only=True)]
            _expand_merges(ws, grid)
            yield ws.title, grid
    finally:
        wb.close()


def _expand_merges(ws, grid) -> None:
    """Repeat each merged cell's value across its whole range.

    A merged tier-1 header like 'Baseline' spanning E4:H4 becomes the value of
    every column in that span, so multi-tier headers can be joined per column.
    """
    for rng in list(ws.merged_cells.ranges):
        r0, c0, r1, c1 = rng.min_row, rng.min_col, rng.max_row, rng.max_col
        if r0 - 1 >= len(grid) or c0 - 1 >= len(grid[r0 - 1]):
            continue
        value = grid[r0 - 1][c0 - 1]
        if value is None:
            continue
        for r in range(r0 - 1, min(r1, len(grid))):
            for c in range(c0 - 1, min(c1, len(grid[r]))):
                if grid[r][c] is None:
                    grid[r][c] = value


class ExcelParser:
    formats = [FileFormat.XLSX, FileFormat.XLS]

    def parse(self, filename: str, data: bytes,
              sheets: Optional[List[str]] = None,
              first_table_only: bool = False) -> ExtractionResult:
        if not _OPENPYXL and not filename.lower().endswith(".xls"):
            raise ParserError("openpyxl is required to read .xlsx files.")
        warnings: List[ParseWarning] = []
        tables: List[FinancialTable] = []
        selected = set(sheets) if sheets else None

        try:
            grids = list(_sheet_grids(filename, data))
        except Exception as exc:  # noqa: BLE001
            raise ParserError(f"Could not read Excel file {filename}: {exc}") from exc

        for sheet_name, grid in grids:
            if selected is not None and sheet_name not in selected:
                continue
            found = tables_from_grid(grid, filename, sheet_name)
            if first_table_only and found:
                found = found[:1]
            if not found:
                warnings.append(ParseWarning(
                    WarningCode.NO_TABLES_FOUND,
                    f"No table detected on sheet '{sheet_name}' of {filename}.",
                    Severity.INFO, {"sheet": sheet_name}))
            tables.extend(found)

        fmt = FileFormat.from_filename(filename)
        doc = FinancialDocument(filename, fmt, tables,
                                metadata={"sheets": [n for n, _ in grids]})
        return ExtractionResult(doc, warnings,
                                {"tables": len(tables), "rows": doc.record_count})
