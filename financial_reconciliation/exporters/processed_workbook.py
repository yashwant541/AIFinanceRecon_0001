"""Write a processed document as a multi-sheet workbook matching the reference
layout: an Index sheet, then one sheet per table with a title row, a blank row,
a header row of [dimensions..., Value], and one row per data point.
"""
from __future__ import annotations

import io
import re
from typing import Dict

import openpyxl
from openpyxl.styles import Font

from .workbook import clean_cell

_SAFE = re.compile(r"[\\/*?:\[\]]")


def _sheet_name(name: str, used: set) -> str:
    nm = _SAFE.sub(" ", name).strip()[:31] or "Table"
    base, i = nm, 1
    while nm.lower() in used:
        i += 1
        nm = f"{base[:28]} {i}"
    used.add(nm.lower())
    return nm


def write_processed_workbook(processed: Dict) -> bytes:
    wb = openpyxl.Workbook()
    index = wb.active
    index.title = "Index"
    index["A1"] = f"{processed.get('source', 'Processed document')} — Extracted Financial Tables"
    index["A1"].font = Font(bold=True, size=13)
    index["A2"] = f"Layout: {processed.get('layout', 'long')} · format: {processed.get('format', '')}"
    index["A4"] = "#"; index["B4"] = "Sheet Name"
    index["A4"].font = index["B4"].font = Font(bold=True)

    used = {"index"}
    for i, t in enumerate(processed.get("tables", []), start=1):
        sname = _sheet_name(t.get("name", f"Table {i}"), used)
        index.cell(row=4 + i, column=1, value=i)
        index.cell(row=4 + i, column=2, value=sname)
        ws = wb.create_sheet(sname)
        ws["A1"] = t.get("name", sname)
        ws["A1"].font = Font(bold=True, size=12)
        cols = t.get("columns") or (list(t["rows"][0].keys()) if t.get("rows") else [])
        for j, c in enumerate(cols):
            cell = ws.cell(row=3, column=1 + j, value=clean_cell(c))
            cell.font = Font(bold=True)
        for r_idx, row in enumerate(t.get("rows", []), start=4):
            for j, c in enumerate(cols):
                ws.cell(row=r_idx, column=1 + j, value=clean_cell(row.get(c)))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
